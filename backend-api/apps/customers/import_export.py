"""
Import/Export utilities for Customer Master
Supports CSV, Excel, and PDF formats
"""
import csv
import logging
from io import BytesIO, StringIO
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.http import HttpResponse
from django.utils import timezone

from .models import CustomerMaster, KAMMaster
from .utils import generate_customer_number

logger = logging.getLogger(__name__)


class CustomerExporter:
    """Export customer data to various formats"""
    
    @staticmethod
    def export_to_csv(queryset):
        """
        Export customers to CSV format
        Returns: HttpResponse with CSV file
        """
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="customers_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        
        writer = csv.writer(response)
        
        # Header row
        headers = [
            'Customer ID',
            'Customer Name',
            'Company Name',
            'Email',
            'Phone',
            'Address',
            'Customer Type',
            'KAM ID',
            'Customer Number',
            'Total Clients',
            'Total Active Clients',
            'Free Giveaway Clients',
            'Default Percentage Share',
            'Contact Person',
            'Status',
            'Last Bill Date',
            'Is Active',
            'Created At',
            'Updated At',
        ]
        writer.writerow(headers)
        
        # Data rows
        for customer in queryset:
            writer.writerow([
                customer.id,
                customer.customer_name,
                customer.company_name or '',
                customer.email,
                customer.phone or '',
                customer.address,
                customer.get_customer_type_display(),
                customer.kam_id.id if customer.kam_id else '',
                customer.customer_number,
                customer.total_client,
                customer.total_active_client,
                customer.free_giveaway_client,
                customer.default_percentage_share,
                customer.contact_person or '',
                customer.get_status_display(),
                customer.last_bill_invoice_date or '',
                'Yes' if customer.is_active else 'No',
                customer.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                customer.updated_at.strftime('%Y-%m-%d %H:%M:%S'),
            ])
        
        return response
    
    @staticmethod
    def export_to_excel(queryset):
        """
        Export customers to Excel format with formatting
        Returns: HttpResponse with Excel file
        """
        # Create DataFrame
        data = []
        for customer in queryset:
            data.append({
                'Customer ID': customer.id,
                'Customer Name': customer.customer_name,
                'Company Name': customer.company_name or '',
                'Email': customer.email,
                'Phone': customer.phone or '',
                'Address': customer.address,
                'Customer Type': customer.get_customer_type_display(),
                'KAM ID': customer.kam_id.id if customer.kam_id else '',
                'Customer Number': customer.customer_number,
                'Total Clients': customer.total_client,
                'Total Active Clients': customer.total_active_client,
                'Free Giveaway Clients': customer.free_giveaway_client,
                'Default % Share': float(customer.default_percentage_share),
                'Contact Person': customer.contact_person or '',
                'Status': customer.get_status_display(),
                'Last Bill Date': customer.last_bill_invoice_date or '',
                'Is Active': 'Yes' if customer.is_active else 'No',
                'Created At': customer.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                'Updated At': customer.updated_at.strftime('%Y-%m-%d %H:%M:%S'),
            })
        
        df = pd.DataFrame(data)
        
        # Create Excel workbook with formatting
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Customers', index=False)
            
            # Get worksheet
            worksheet = writer.sheets['Customers']
            
            # Format header row
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Set column widths
            column_widths = {
                'A': 12,  # Customer ID
                'B': 20,  # Customer Name
                'C': 20,  # Company Name
                'D': 25,  # Email
                'E': 15,  # Phone
                'F': 20,  # Address
                'G': 15,  # Customer Type
                'H': 12,  # KAM ID
                'I': 18,  # Customer Number
                'J': 15,  # Total Clients
                'K': 18,  # Total Active Clients
                'L': 18,  # Free Giveaway Clients
                'M': 15,  # Default % Share
                'N': 15,  # Contact Person
                'O': 12,  # Status
                'P': 15,  # Last Bill Date
                'Q': 10,  # Is Active
                'R': 20,  # Created At
                'S': 20,  # Updated At
            }
            
            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width
            
            # Format data rows
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical='center')
        
        output.seek(0)
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="customers_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        return response


class CustomerImporter:
    """Import customer data from CSV or Excel formats"""
    
    @staticmethod
    def validate_required_fields(row_data, row_number):
        """
        Validate required fields in import data
        Returns: (is_valid, error_messages)
        """
        errors = []
        required_fields = ['Customer Name', 'Email', 'Address', 'Customer Type']
        
        for field in required_fields:
            if not row_data.get(field, '').strip():
                errors.append(f"Row {row_number}: Missing required field '{field}'")
        
        # Validate email format
        email = row_data.get('Email', '').strip()
        if email and '@' not in email:
            errors.append(f"Row {row_number}: Invalid email format '{email}'")
        
        # Validate customer type
        valid_types = dict(CustomerMaster.CUSTOMER_TYPE_CHOICES)
        customer_type = row_data.get('Customer Type', '').strip().lower()
        if customer_type and customer_type not in valid_types:
            errors.append(f"Row {row_number}: Invalid customer type '{customer_type}'. Must be one of: {', '.join(valid_types.keys())}")
        
        return len(errors) == 0, errors
    
    @staticmethod
    def import_from_csv(file_content):
        """
        Import customers from CSV file
        Returns: (success_count, error_messages, created_customers)
        """
        success_count = 0
        error_messages = []
        created_customers = []
        
        try:
            # Read CSV content
            csv_file = StringIO(file_content.decode('utf-8'))
            reader = csv.DictReader(csv_file)
            
            if not reader.fieldnames:
                return 0, ['Invalid CSV file: No headers found'], []
            
            for row_number, row in enumerate(reader, start=2):
                try:
                    # Validate row
                    is_valid, validation_errors = CustomerImporter.validate_required_fields(row, row_number)
                    if not is_valid:
                        error_messages.extend(validation_errors)
                        continue
                    
                    # Get KAM by ID
                    kam = None
                    kam_id = row.get('KAM ID', '').strip()
                    if kam_id:
                        try:
                            kam = KAMMaster.objects.get(id=int(kam_id))
                        except (ValueError, KAMMaster.DoesNotExist):
                            error_messages.append(f"Row {row_number}: KAM with ID '{kam_id}' not found")
                            continue
                    
                    # Check if customer already exists
                    email = row.get('Email', '').strip().lower()
                    if CustomerMaster.objects.filter(email=email).exists():
                        error_messages.append(f"Row {row_number}: Customer with email '{email}' already exists")
                        continue
                    
                    # Create customer
                    customer = CustomerMaster.objects.create(
                        customer_name=row.get('Customer Name', '').strip(),
                        company_name=row.get('Company Name', '').strip() or None,
                        email=email,
                        phone=row.get('Phone', '').strip() or '',
                        address=row.get('Address', '').strip(),
                        customer_type=row.get('Customer Type', '').strip().lower(),
                        kam_id=kam,
                        contact_person=row.get('Contact Person', '').strip() or '',
                        total_client=int(row.get('Total Clients', 0)) or 0,
                        total_active_client=int(row.get('Total Active Clients', 0)) or 0,
                        free_giveaway_client=int(row.get('Free Giveaway Clients', 0)) or 0,
                        default_percentage_share=float(row.get('Default Percentage Share', 0)) or 0,
                        status=row.get('Status', 'active').lower(),
                    )
                    
                    # Auto-generate customer number
                    customer.customer_number = generate_customer_number(customer.customer_name, customer.id)
                    customer.save()
                    
                    # Prepare response with KAM details
                    customer_data = {
                        'id': customer.id,
                        'customer_name': customer.customer_name,
                        'customer_number': customer.customer_number,
                        'email': customer.email,
                    }
                    if customer.kam_id:
                        customer_data['kam'] = {
                            'id': customer.kam_id.id,
                            'name': customer.kam_id.kam_name,
                            'email': customer.kam_id.email,
                        }
                    created_customers.append(customer_data)
                    success_count += 1
                    
                except ValueError as e:
                    error_messages.append(f"Row {row_number}: Invalid data format - {str(e)}")
                except Exception as e:
                    error_messages.append(f"Row {row_number}: Error creating customer - {str(e)}")
        
        except Exception as e:
            error_messages.append(f"Error reading CSV file: {str(e)}")
        
        return success_count, error_messages, created_customers
    
    @staticmethod
    def import_from_excel(file_content):
        """
        Import customers from Excel file
        Returns: (success_count, error_messages, created_customers)
        """
        success_count = 0
        error_messages = []
        created_customers = []
        
        try:
            # Read Excel file
            df = pd.read_excel(BytesIO(file_content))
            
            if df.empty:
                return 0, ['Excel file is empty'], []
            
            for row_number, row in df.iterrows():
                try:
                    # Convert row to dictionary
                    row_data = row.to_dict()
                    
                    # Handle NaN values
                    row_data = {k: ('' if pd.isna(v) else str(v)) for k, v in row_data.items()}
                    
                    # Validate row
                    is_valid, validation_errors = CustomerImporter.validate_required_fields(row_data, row_number + 2)
                    if not is_valid:
                        error_messages.extend(validation_errors)
                        continue
                    
                    # Get KAM by ID
                    kam = None
                    kam_id = row_data.get('KAM ID', '').strip()
                    if kam_id:
                        try:
                            kam = KAMMaster.objects.get(id=int(kam_id))
                        except (ValueError, KAMMaster.DoesNotExist):
                            error_messages.append(f"Row {row_number + 2}: KAM with ID '{kam_id}' not found")
                            continue
                    
                    # Check if customer already exists
                    email = row_data.get('Email', '').strip().lower()
                    if CustomerMaster.objects.filter(email=email).exists():
                        error_messages.append(f"Row {row_number + 2}: Customer with email '{email}' already exists")
                        continue
                    
                    # Create customer
                    customer = CustomerMaster.objects.create(
                        customer_name=row_data.get('Customer Name', '').strip(),
                        company_name=row_data.get('Company Name', '').strip() or None,
                        email=email,
                        phone=row_data.get('Phone', '').strip() or '',
                        address=row_data.get('Address', '').strip(),
                        customer_type=row_data.get('Customer Type', '').strip().lower(),
                        kam_id=kam,
                        contact_person=row_data.get('Contact Person', '').strip() or '',
                        total_client=int(float(row_data.get('Total Clients', 0)) or 0),
                        total_active_client=int(float(row_data.get('Total Active Clients', 0)) or 0),
                        free_giveaway_client=int(float(row_data.get('Free Giveaway Clients', 0)) or 0),
                        default_percentage_share=float(row_data.get('Default % Share', 0)) or 0,
                        status=row_data.get('Status', 'active').lower(),
                    )
                    
                    # Auto-generate customer number
                    customer.customer_number = generate_customer_number(customer.customer_name, customer.id)
                    customer.save()
                    
                    # Prepare response with KAM details
                    customer_data = {
                        'id': customer.id,
                        'customer_name': customer.customer_name,
                        'customer_number': customer.customer_number,
                        'email': customer.email,
                    }
                    if customer.kam_id:
                        customer_data['kam'] = {
                            'id': customer.kam_id.id,
                            'name': customer.kam_id.kam_name,
                            'email': customer.kam_id.email,
                        }
                    created_customers.append(customer_data)
                    success_count += 1
                    
                except ValueError as e:
                    error_messages.append(f"Row {row_number + 2}: Invalid data format - {str(e)}")
                except Exception as e:
                    error_messages.append(f"Row {row_number + 2}: Error creating customer - {str(e)}")
        
        except Exception as e:
            error_messages.append(f"Error reading Excel file: {str(e)}")
        
        return success_count, error_messages, created_customers
